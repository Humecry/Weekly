#! /usr/bin/env python3
# -*- coding:utf-8 -*-

# @Date    : 2018-06-11 20:07:43
# @Author  : Hume (102734075@qq.com)
# @Link    : https://humecry.wordpress.com/
# @Version : 1.2
# @Description：获取京东到家与美团外卖各门店上周数据并添加进Excel

import requests
import sys, json
import time, datetime
import pandas as pd
from datetime import timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.styles import Font, colors, Alignment, Border, Side, numbers, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
# 引入配置文件
from conf import *
# 引入公共函数
from common import *

# 当前时间
now = datetime.datetime.now()
# 上周的第一天周日，最后一天周六
last_week_start = now - timedelta(days=7+now.isoweekday())
last_week_end = now - timedelta(days=1+now.isoweekday())
'''
获取京东到家上周数据
'''
def getJD(headers, shops, process_bar):
	dfs = []
	# 添加上周京东到家数据
	for shop in shops.values():
		timeDate = last_week_start.strftime("%Y-%m-%d") + "~" + last_week_end.strftime("%Y-%m-%d")
		# 2018年
		# yearWeek = time.strftime("%Y", time.localtime(time.time()-7*24*60*60)) + '年' + str(int(time.strftime("%W", time.localtime(time.time()-7*24*60*60)))) + '周'
		# 2019年
		yearWeek = time.strftime("%Y", time.localtime(time.time())) + '年' + str(int(time.strftime("%W", time.localtime(time.time())))) + '周'
		params = {
		    "venderId": "320695",
		    "timeRangeType": "2",
		    "weekId": yearWeek,
		    "dateWeek": yearWeek,
		    "shopIdListStr": shop
		}
		try:
			# 获取主要数据
			url = 'https://dc-store.jd.com/operation/queryShopOperationData'
			response = requests.get(url, headers=headers, params=params, timeout=10)
			json = response.json()
			data = json['operationDataResponseDTOs'][0]
		except:
			if response.text.find('loginContent') != -1:
				print('京东到家报错:获取数据失败,cookie过期,请更换cookie!')
			else:
				print(response.text + '\n京东到家报错:以上是京东到家网站返回的信息')
			return False

		process_bar.show_process()

		try:
			# 获取环比
			url2 = 'https://dc-store.jd.com/operation/queryData'
			response2 = requests.get(url2, headers=headers, params=params, timeout=10)
			json2 = response2.json()
			row = [timeDate, data['shopName'], data['browseCnt'], data['totalVisitCnt'], data['validOrderCnt'], data['takeRate']/100, data['orderTotalAmtz'], data['perTicketSales'], 0 if json2['validOrderCountRelativeRatio']=='--' else json2['validOrderCountRelativeRatioSign']*float(json2['validOrderCountRelativeRatio'])/100]
		except:
			if response2.text.find('loginContent') != -1:
				print('京东到家报错:获取数据失败,cookie过期,请更换cookie!')
			else:
				print(response2.text + '\n京东到家报错:以上是京东到家网站返回的信息')
			return False
		dfs.append(pd.DataFrame([row], columns=['日期','门店', '浏览量', '访客数', '有效订单数', '转化率', 'GMV成交额', '客单价', '环比']))

		process_bar.show_process()

	df1 = dfs.pop()
	for value in dfs:
		df1 = pd.concat([df1, value], ignore_index=True)
	return df1
'''
获取美团外卖华森店上周数据
'''
def getMeiTuan(headers, process_bar):
	timeDate = last_week_start.strftime("%Y-%m-%d") + "~" + last_week_end.strftime("%Y-%m-%d")
	params = {
	    'wmPoiId': '3578168',
	    'beginTime': last_week_start.strftime("%Y%m%d"),
	    'endTime': last_week_end.strftime("%Y%m%d"),
	}
	try:
		# 获取主要数据
		url = 'http://waimaieapp.meituan.com/bizdata/businessStatisticsV3/single/hisOverview'
		response = requests.get(url, headers=headers, params=params, timeout=10)
		json = response.json()
		data = json['data']
	except:
		if response.text.find('登录信息不完整，请重新登录') != -1:
			print('美团外卖报错:获取数据失败,cookie过期,请更换cookie!')
		else:
			print(response.text + '\n美团外卖报错:以上是美团外卖网站返回的信息')
		return False
	params2 = {
		'wmPoiId': '3578168',
		'recentDays': '7'
	}

	process_bar.show_process()

	try:
		# # 获取环比
		url2 = 'http://waimaieapp.meituan.com/bizdata/flowanalysisV2/flow/overview'
		response2 = requests.get(url2, headers=headers, params=params2, timeout=10)
		json2 = response2.json()
		data2 = json2['data']['flowOverviewInfo']
		row = [timeDate, '美团外卖-华森店', data2['exposureNum'], data2['visitNum'], data['effectiveOrders'], data['effectiveOrders']/data2['visitNum'], data['turnover'], data['turnover']/data['effectiveOrders'], (data['effectiveOrders']-data['effectiveOrdersLastPeriod'])/data['effectiveOrdersLastPeriod']]
	except:
		if response2.text.find('登录信息不完整，请重新登录') != -1:
			print('美团外卖报错:获取数据失败,cookie过期,请更换cookie!')
		else:
			print(response2.text + '\n美团外卖报错:以上是美团外卖网站返回的信息')
		return False
	df = (pd.DataFrame([row], columns=['日期','门店', '浏览量', '访客数', '有效订单数', '转化率', 'GMV成交额', '客单价', '环比']))

	process_bar.show_process()

	return df
# 将字典格式化为json字符串
def echo(dic):
	jsonString = json.dumps(dic, ensure_ascii=False, indent=4)
	print(jsonString)
	return jsonString

def main(process_bar):
	# 要修改的Excel文件
	filePath = PATH + '网销每周数据.xlsx' 
	# 获取Excel数据
	try:
		wb = load_workbook(filePath)
	except FileNotFoundError:
		print('网销数据报错:在"' + PATH + '"中找不到"网销每周数据.xlsx"文件!')
		return False
	sheet = wb['网销周数据']
	data = sheet.values
	cols = next(data)[:] # 获取表头
	ExcelDf = pd.DataFrame(data, columns=cols)
	# 京东到家
	JDheaders = {
		'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.117 Safari/537.36',
		'Cookie': netSales['JDcookie'],
	}
	JDdf = getJD(JDheaders, JDshops, process_bar)
	if isinstance(JDdf, bool):
		return False
	# 美团外卖华森店
	# MTheaders = {
	# 	'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.117 Safari/537.36',
	# 	'Cookie': netSales['MTcookie'],
	# }
	# MTdf = getMeiTuan(MTheaders, process_bar)
	# if isinstance(MTdf, bool):
	# 	return False
	# 拼接数据
	# df = pd.concat([JDdf, MTdf], ignore_index=True)

	df = pd.concat([ExcelDf, JDdf], ignore_index=True)
	df.drop_duplicates(['日期', '门店'],inplace=True)
	# 使用pandas进行排序
	df.sort_values(['门店', '日期'], inplace=True)
	# 删除原Excel表格
	wb.remove(sheet)
	# 创建新表
	sheet = wb.create_sheet("网销周数据")
	for r in dataframe_to_rows(df, index=False):
	    sheet.append(r)
	tab = Table(displayName="Table1", ref="A1:I"+str(len(df)+1))
	style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
	                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
	tab.tableStyleInfo = style
	sheet.add_table(tab)
	# 冻结窗格
	sheet.freeze_panes = 'A2'
	# 设置单元格列宽
	sheet.column_dimensions['A'].width = 33
	sheet.column_dimensions['B'].width = 20
	sheet.column_dimensions['C'].width = 15
	sheet.column_dimensions['D'].width = 15
	sheet.column_dimensions['E'].width = 15
	sheet.column_dimensions['F'].width = 15
	sheet.column_dimensions['G'].width = 15
	sheet.column_dimensions['H'].width = 15
	sheet.column_dimensions['I'].width = 15
	# 单元格边框
	border = Border(left=Side(style='thin',color='FF000000'),right=Side(style='thin',color='FF000000'),top=Side(style='thin',color='FF000000'),bottom=Side(style='thin',color='FF000000'),diagonal=Side(style='thin',color='FF000000'),diagonal_direction=0,outline=Side(style='thin',color='FF000000'),vertical=Side(style='thin',color='FF000000'),horizontal=Side(style='thin',color='FF000000'))
	# 所有单元格设置字体
	for i in range(sheet.max_row):
		for k in range(sheet.max_column):
			cell = sheet.cell(row=i+1, column=k+1)
			if k==0 or i==0:
				# 设置首行首列加粗
				cell.font = Font(name="微软雅黑", bold=True, size=13)
				# 垂直居中和水平居中
				cell.alignment = Alignment(horizontal='center', vertical='center')
			else:
				# 设置百分比数据格式
				if k==5 or k==8:
					cell.number_format = numbers.FORMAT_PERCENTAGE_00
				if k==6:
					cell.number_format = "0.00"
				cell.font = Font(name="微软雅黑", size=13)
				if k==7:
					cell.number_format = "0.00"
				cell.font = Font(name="微软雅黑", size=13)
			# 设置单元格边框
			cell.border = border

	# 条件格式
	redFill = PatternFill(start_color='FF0040', end_color='FF0040', fill_type='solid')
	greenFill = PatternFill(start_color='01DF3A', end_color='01DF3A', fill_type='solid')
	red_text = Font(color="9C0006")
	red_fill = PatternFill(bgColor="FFC7CE")
	dxf = DifferentialStyle(font=red_text, fill=red_fill)
	rule = Rule(type="containsText", operator="containsText", text=last_week_start.strftime("%Y-%m-%d"), dxf=dxf)
	sheet.conditional_formatting.add('I2:I'+str(len(df)+1), CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=redFill))
	sheet.conditional_formatting.add('I2:I'+str(len(df)+1), CellIsRule(operator='greaterThanOrEqual', formula=['0'], stopIfTrue=True, fill=greenFill))
	sheet.conditional_formatting.add('A2:A'+str(len(df)+1), rule)

	# 有效订单表
	try:
		sheet2 = wb['有效订单数']
		wb.remove(sheet2)
	except:
		pass
	df2 = df.pivot(index='日期', columns='门店', values='有效订单数')
	df2.reset_index(inplace=True)
	# df2.columns = ['日期', '京东华森店', '塔埔小店', '塔埔店', '绿苑店', '美团华森店']
	sheet2 = wb.create_sheet('有效订单数')
	for r in dataframe_to_rows(df2, index=False):
		sheet2.append(r)
	tab = Table(displayName="Table1", ref="A1:"+chr(64+df2.columns.size)+str(df2.index.size+1))
	style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
	                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
	tab.tableStyleInfo = style
	sheet2.add_table(tab)
	# 冻结窗格
	sheet2.freeze_panes = 'A2'
	# 设置单元格列宽
	sheet2.column_dimensions['A'].width = 24
	for i in range(df2.columns.size-1):
		sheet2.column_dimensions[chr(66+i)].width = 18
	# 作图
	chart = LineChart()
	chart.title = "有效订单数统计"
	chart.style = 2
	chart.x_axis.title = '日期'
	chart.y_axis.title = '有效订单数'
	chart.height = 12
	chart.width = 19
	cats = Reference(sheet2, min_col=1, min_row=2, max_row=df2.index.size+1)
	data = Reference(sheet2, min_col=2, min_row=1, max_col=df2.columns.size, max_row=df2.index.size+1)
	chart.add_data(data, titles_from_data=True)
	chart.set_categories(cats)
	sheet2.add_chart(chart, chr(66+df2.columns.size)+"2")
	# 保存Excel
	wb.save(filename=filePath)
	return filePath

if __name__ == '__main__':
	# 进度条
	max_steps = len(JDshops) * 2
	process_bar = ShowProcess(max_steps, 'OK')
	main(process_bar)